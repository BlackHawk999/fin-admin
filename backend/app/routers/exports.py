from datetime import date
from io import BytesIO
from typing import Annotated

import pytz
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import User, Expense, Owner
from ..auth import get_current_user
from ..config import get_settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/exports", tags=["exports"])
settings = get_settings()
TZ = pytz.timezone(settings.timezone)


@router.get("/expenses.xlsx")
def export_expenses_xlsx(
    db: Annotated[Session, Depends(get_db)],
    _: Annotated[User, Depends(get_current_user)],
    date_from: date = Query(..., description="Start date"),
    date_to: date = Query(..., description="End date"),
    category: str | None = None,
    owner_id: int | None = None,
):
    q = db.query(Expense).filter(
        Expense.date >= date_from,
        Expense.date <= date_to,
    )
    if category:
        q = q.filter(Expense.category.ilike(f"%{category}%"))
    if owner_id is not None:
        q = q.filter(Expense.owner_id == owner_id)
    expenses = q.order_by(Expense.date).all()

    owners = {o.id: o for o in db.query(Owner).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    header_font = Font(bold=True)
    ws["A1"] = "Отчёт по расходам"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Период: {date_from} — {date_to}"
    ws["A2"].font = header_font

    row = 4
    ws.cell(row=row, column=1, value="Date")
    ws.cell(row=row, column=2, value="Amount (UZS)")
    ws.cell(row=row, column=3, value="Category")
    ws.cell(row=row, column=4, value="Owner")
    ws.cell(row=row, column=5, value="Comment")
    for c in range(1, 6):
        ws.cell(row=row, column=c).font = header_font
    row += 1

    total = 0
    for e in expenses:
        owner_name = ""
        if e.owner_id and e.owner_id in owners:
            owner_name = owners[e.owner_id].name
        ws.cell(row=row, column=1, value=str(e.date))
        ws.cell(row=row, column=2, value=int(e.amount_uzs))
        ws.cell(row=row, column=3, value=e.category or "")
        ws.cell(row=row, column=4, value=owner_name)
        ws.cell(row=row, column=5, value=(e.comment or ""))
        total += int(e.amount_uzs)
        row += 1

    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=total)
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=2).font = header_font

    for c in range(1, 6):
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(str(ws.cell(row=4, column=c).value or "")) + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"expenses_{date_from}_to_{date_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
